from __future__ import annotations

import calendar
import hashlib
import json
import math
import os
import re
import shutil
import tempfile
from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.workbook.properties import CalcProperties
from openpyxl.worksheet.datavalidation import DataValidation


WORKBOOK_FILENAME = "ESTOQUE SICRONIZADO.xlsx"
CURRENT_SHEET_TITLE = "ESTOQUE ATUAL"
EXCEL_ONLINE_URL = (
    "https://1drv.ms/x/c/74b99486c97e7a7a/"
    "IQBP69FZZxvpSJd-n362-ZLIAeb3FLFulV5v0Ie3uPDxJWQ?e=9jkgpC"
)
MONTH_NAMES = (
    "",
    "JANEIRO",
    "FEVEREIRO",
    "MARÇO",
    "ABRIL",
    "MAIO",
    "JUNHO",
    "JULHO",
    "AGOSTO",
    "SETEMBRO",
    "OUTUBRO",
    "NOVEMBRO",
    "DEZEMBRO",
)
MONTH_BY_NAME = {name: number for number, name in enumerate(MONTH_NAMES) if name}
HEADERS = ("PRODUTO", "ESTOQUE DO SISTEMA", "CONTAGEM", "DIFERENÇA", "ESTOQUE FINAL")
HIDDEN_HEADERS = ("ID DO PRODUTO", "MOVIMENTOS APÓS CONTAGEM", "MÊS")
CURRENT_HEADERS = ("PRODUTO", "ESTOQUE ATUAL")


class ExcelSyncError(RuntimeError):
    pass


@dataclass(frozen=True)
class WorkbookCount:
    month: str
    product_id: int
    quantity: float


def default_workbook_path() -> Path | None:
    roots: list[Path] = []
    for key in ("OneDrive", "OneDriveConsumer", "OneDriveCommercial"):
        value = os.environ.get(key)
        if value:
            roots.append(Path(value))
    roots.append(Path.home() / "OneDrive")
    checked: set[Path] = set()
    for root in roots:
        candidate = root / WORKBOOK_FILENAME
        if candidate in checked:
            continue
        checked.add(candidate)
        if candidate.is_file():
            return candidate
    return None


def month_title(month: str) -> str:
    match = re.fullmatch(r"(\d{4})-(\d{2})", month)
    if not match:
        raise ValueError("Mês inválido para a planilha.")
    year, number = int(match.group(1)), int(match.group(2))
    if number not in range(1, 13):
        raise ValueError("Mês inválido para a planilha.")
    return f"{MONTH_NAMES[number]} {year}"


def month_from_title(title: str) -> str | None:
    match = re.fullmatch(r"\s*([A-ZÇ]+)\s+(\d{4})\s*", title.upper())
    if not match or match.group(1) not in MONTH_BY_NAME:
        return None
    return f"{int(match.group(2)):04d}-{MONTH_BY_NAME[match.group(1)]:02d}"


def month_last_day(month: str) -> str:
    year, number = (int(part) for part in month.split("-"))
    return f"{year:04d}-{number:02d}-{calendar.monthrange(year, number)[1]:02d}"


def combined_product_name(product: dict) -> str:
    return " ".join(
        str(product.get(key) or "").strip()
        for key in ("group_name", "name", "variant")
        if str(product.get(key) or "").strip()
    ).upper()


def _count_number(value: object) -> float | None:
    if value in (None, "") or isinstance(value, bool):
        return None
    try:
        number = float(str(value).strip().replace(",", "."))
    except (TypeError, ValueError):
        return None
    if not math.isfinite(number) or number < 0:
        return None
    return number


def workbook_data_fingerprint(months: list[dict]) -> str:
    serialized = json.dumps(
        months,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
        default=str,
    )
    return hashlib.sha256(serialized.encode("utf-8")).hexdigest()


class MonthlyStockWorkbook:
    def __init__(self, path: Path | None = None):
        self.path = Path(path) if path else default_workbook_path()
        if self.path is None:
            raise ExcelSyncError(
                f"O arquivo {WORKBOOK_FILENAME} não foi encontrado na pasta do OneDrive."
            )

    def _load(self):
        try:
            return load_workbook(self.path, data_only=False)
        except PermissionError as error:
            raise ExcelSyncError("A planilha está ocupada. Feche o Excel e tente novamente.") from error
        except (OSError, ValueError) as error:
            raise ExcelSyncError("Não foi possível abrir a planilha sincronizada.") from error

    def read_counts(self) -> list[WorkbookCount]:
        workbook = self._load()
        counts: list[WorkbookCount] = []
        for sheet in workbook.worksheets:
            month = month_from_title(sheet.title)
            if not month:
                continue
            for row in range(2, sheet.max_row + 1):
                product_id = sheet.cell(row, 6).value
                quantity = _count_number(sheet.cell(row, 3).value)
                if quantity is None:
                    continue
                try:
                    counts.append(WorkbookCount(month, int(product_id), quantity))
                except (TypeError, ValueError):
                    continue
        workbook.close()
        return counts

    def write(self, months: list[dict]) -> dict:
        workbook = self._load() if self.path.is_file() else Workbook()
        current_month = next((item for item in months if item.get("is_current")), None)
        current_rows = list((current_month or {}).get("rows") or [])
        if CURRENT_SHEET_TITLE in workbook.sheetnames:
            current_sheet = workbook[CURRENT_SHEET_TITLE]
            self._reset_sheet(current_sheet)
        else:
            current_sheet = workbook.create_sheet(CURRENT_SHEET_TITLE, 0)
        self._move_sheet(workbook, current_sheet, 0)
        self._write_current(current_sheet, current_rows)

        written: list[str] = [CURRENT_SHEET_TITLE]
        for position, month_data in enumerate(months, start=1):
            key = str(month_data["month"])
            title = month_title(key)
            existing = workbook[title] if title in workbook.sheetnames else None
            if existing is None:
                sheet = workbook.create_sheet(title, position)
            else:
                sheet = existing
                self._reset_sheet(sheet)
            self._move_sheet(workbook, sheet, position)
            self._write_month(sheet, month_data)
            written.append(title)

        for placeholder_title in ("Planilha1", "Sheet"):
            if placeholder_title in workbook.sheetnames and len(workbook.sheetnames) > 1:
                placeholder = workbook[placeholder_title]
                if placeholder.max_row == 1 and placeholder.max_column == 1 and placeholder["A1"].value is None:
                    workbook.remove(placeholder)

        for sheet in workbook.worksheets:
            sheet.sheet_view.tabSelected = False
        workbook.active = 0
        workbook[CURRENT_SHEET_TITLE].sheet_view.tabSelected = True
        if workbook.views:
            workbook.views[0].activeTab = 0
            workbook.views[0].firstSheet = 0

        if workbook.calculation is None:
            workbook.calculation = CalcProperties()
        workbook.calculation.fullCalcOnLoad = True
        workbook.calculation.forceFullCalc = True
        temporary_path: Path | None = None
        try:
            with tempfile.NamedTemporaryFile(
                prefix=f".{self.path.stem}-", suffix=".xlsx", dir=self.path.parent, delete=False
            ) as temporary:
                temporary_path = Path(temporary.name)
            workbook.save(temporary_path)
            workbook.close()
            if self.path.exists():
                # Mantém a identidade do arquivo para o cliente do OneDrive
                # reconhecer a alteração imediatamente, sem criar um novo item.
                with temporary_path.open("rb") as source, self.path.open("r+b") as destination:
                    shutil.copyfileobj(source, destination)
                    destination.truncate()
                    destination.flush()
                    os.fsync(destination.fileno())
            else:
                os.replace(temporary_path, self.path)
        except PermissionError as error:
            raise ExcelSyncError("A planilha está ocupada. Feche o Excel e tente novamente.") from error
        except OSError as error:
            raise ExcelSyncError("Não foi possível salvar a planilha no OneDrive.") from error
        finally:
            if temporary_path and temporary_path.exists():
                temporary_path.unlink(missing_ok=True)
        return {"path": str(self.path), "sheets": written}

    @staticmethod
    def _reset_sheet(sheet) -> None:
        for merged in list(sheet.merged_cells.ranges):
            sheet.unmerge_cells(str(merged))
        if sheet.max_row:
            sheet.delete_rows(1, sheet.max_row)
        sheet.conditional_formatting._cf_rules.clear()
        sheet.data_validations.dataValidation = []
        sheet.auto_filter.ref = None

    @staticmethod
    def _move_sheet(workbook, sheet, position: int) -> None:
        current_position = workbook.index(sheet)
        if current_position != position:
            workbook.move_sheet(sheet, offset=position - current_position)

    @staticmethod
    def _write_current(sheet, rows: list[dict]) -> None:
        sheet.sheet_view.showGridLines = False
        sheet.freeze_panes = "A2"
        sheet.sheet_properties.tabColor = "00A6A6"
        sheet.row_dimensions[1].height = 28
        sheet.column_dimensions["A"].width = 39
        sheet.column_dimensions["B"].width = 20

        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(name="Aptos", size=10, bold=True, color="FFFFFF")
        thin_blue = Side(style="thin", color="7F9DB9")
        body_border = Border(left=thin_blue, right=thin_blue, top=thin_blue, bottom=thin_blue)
        sheet.append(CURRENT_HEADERS)
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = body_border

        for index, item in enumerate(rows, start=2):
            final_stock = item.get("final_stock")
            if final_stock is None:
                counted = item.get("counted")
                final_stock = (
                    item.get("system_stock")
                    if counted is None
                    else float(counted) + float(item.get("post_count_delta") or 0)
                )
            sheet.cell(index, 1, str(item["product"]))
            sheet.cell(index, 2, float(final_stock or 0))
            for column in range(1, 3):
                cell = sheet.cell(index, column)
                cell.font = Font(name="Aptos", size=10, bold=column == 1)
                cell.border = body_border
                cell.alignment = Alignment(
                    horizontal="left" if column == 1 else "center", vertical="center"
                )
                cell.protection = Protection(locked=True)
            sheet.cell(index, 1).fill = PatternFill(
                "solid", fgColor="D9EAF2" if index % 2 == 0 else "EAF5FA"
            )
            sheet.cell(index, 2).fill = PatternFill("solid", fgColor="E2F0D9")
            sheet.cell(index, 2).number_format = '#,##0'
            sheet.row_dimensions[index].height = 22

        last_data_row = max(2, len(rows) + 1)
        total_row = len(rows) + 2
        sheet.cell(total_row, 1, "TOTAL")
        sheet.cell(total_row, 2, f"=SUM(B2:B{last_data_row})" if rows else "=0")
        for column in range(1, 3):
            cell = sheet.cell(total_row, column)
            cell.fill = PatternFill("solid", fgColor="D9EAD3")
            cell.font = Font(name="Aptos", size=11, bold=True, color="1F1F1F")
            cell.border = Border(top=Side(style="medium", color="1F4E78"))
            cell.alignment = Alignment(horizontal="left" if column == 1 else "center")
            cell.protection = Protection(locked=True)
        sheet.cell(total_row, 2).number_format = '#,##0'
        sheet.auto_filter.ref = f"A1:B{last_data_row}"

        if rows:
            sheet.conditional_formatting.add(
                f"B2:B{last_data_row}",
                CellIsRule(
                    operator="lessThan",
                    formula=["0"],
                    fill=PatternFill("solid", fgColor="5A0B1A"),
                    font=Font(color="FFFFFF", bold=True),
                ),
            )

        sheet.protection.sheet = True
        sheet.protection.autoFilter = False
        sheet.protection.sort = False
        sheet.protection.selectLockedCells = True
        sheet.protection.selectUnlockedCells = False

    @staticmethod
    def _write_month(sheet, month_data: dict) -> None:
        rows = list(month_data.get("rows") or [])
        sheet.sheet_view.showGridLines = False
        sheet.freeze_panes = "A2"
        sheet.sheet_properties.tabColor = "1F4E78"
        sheet.row_dimensions[1].height = 28
        sheet.column_dimensions["A"].width = 39
        sheet.column_dimensions["B"].width = 20
        sheet.column_dimensions["C"].width = 15
        sheet.column_dimensions["D"].width = 14
        sheet.column_dimensions["E"].width = 17
        for column in ("F", "G", "H"):
            sheet.column_dimensions[column].hidden = True

        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(name="Aptos", size=10, bold=True, color="FFFFFF")
        thin_blue = Side(style="thin", color="7F9DB9")
        body_border = Border(left=thin_blue, right=thin_blue, top=thin_blue, bottom=thin_blue)
        sheet.append([*HEADERS, *HIDDEN_HEADERS])
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = body_border

        for index, item in enumerate(rows, start=2):
            counted = item.get("counted")
            sheet.cell(index, 1, str(item["product"]))
            sheet.cell(index, 2, float(item["system_stock"]))
            if counted is not None:
                sheet.cell(index, 3, float(counted))
            sheet.cell(index, 4, f'=IF(C{index}="","",C{index}-B{index})')
            sheet.cell(index, 5, f'=IF(C{index}="",B{index},C{index}+G{index})')
            sheet.cell(index, 6, int(item["product_id"]))
            sheet.cell(index, 7, float(item.get("post_count_delta") or 0))
            sheet.cell(index, 8, str(month_data["month"]))
            for column in range(1, 9):
                cell = sheet.cell(index, column)
                cell.font = Font(name="Aptos", size=10, bold=column == 1)
                cell.border = body_border
                cell.alignment = Alignment(
                    horizontal="left" if column == 1 else "center", vertical="center"
                )
                cell.protection = Protection(locked=column != 3)
            sheet.cell(index, 1).fill = PatternFill(
                "solid", fgColor="D9EAF2" if index % 2 == 0 else "EAF5FA"
            )
            sheet.cell(index, 2).fill = PatternFill("solid", fgColor="E2F0D9")
            sheet.cell(index, 3).fill = PatternFill("solid", fgColor="FFF2CC")
            sheet.cell(index, 5).fill = PatternFill("solid", fgColor="DDEBF7")
            for column in (2, 3, 4, 5, 7):
                sheet.cell(index, column).number_format = '#,##0'
            sheet.row_dimensions[index].height = 22

        last_data_row = max(2, len(rows) + 1)
        total_row = len(rows) + 2
        sheet.cell(total_row, 1, "TOTAL")
        sheet.cell(total_row, 5, f"=SUM(E2:E{last_data_row})" if rows else "=0")
        for column in range(1, 6):
            cell = sheet.cell(total_row, column)
            cell.fill = PatternFill("solid", fgColor="D9EAD3")
            cell.font = Font(name="Aptos", size=11, bold=True, color="1F1F1F")
            cell.border = Border(top=Side(style="medium", color="1F4E78"))
            cell.alignment = Alignment(horizontal="left" if column == 1 else "center")
            cell.protection = Protection(locked=True)
        sheet.cell(total_row, 5).number_format = '#,##0'
        sheet.auto_filter.ref = f"A1:E{last_data_row}"

        if rows:
            count_validation = DataValidation(
                type="decimal", operator="greaterThanOrEqual", formula1="0", allow_blank=True
            )
            count_validation.error = "Digite uma quantidade igual ou maior que zero."
            count_validation.errorTitle = "Contagem inválida"
            count_validation.prompt = "Informe a quantidade física encontrada."
            count_validation.promptTitle = "Contagem"
            count_validation.showErrorMessage = True
            count_validation.showInputMessage = True
            sheet.add_data_validation(count_validation)
            count_validation.add(f"C2:C{last_data_row}")

            difference_range = f"D2:D{last_data_row}"
            sheet.conditional_formatting.add(
                difference_range,
                CellIsRule(
                    operator="lessThan",
                    formula=["0"],
                    fill=PatternFill("solid", fgColor="F4CCCC"),
                    font=Font(color="9C0006", bold=True),
                ),
            )
            sheet.conditional_formatting.add(
                difference_range,
                CellIsRule(
                    operator="greaterThan",
                    formula=["0"],
                    fill=PatternFill("solid", fgColor="D9EAD3"),
                    font=Font(color="006100", bold=True),
                ),
            )
            sheet.conditional_formatting.add(
                f"B2:B{last_data_row}",
                CellIsRule(
                    operator="lessThan",
                    formula=["0"],
                    fill=PatternFill("solid", fgColor="5A0B1A"),
                    font=Font(color="FFFFFF", bold=True),
                ),
            )

        sheet.protection.sheet = True
        sheet.protection.autoFilter = False
        sheet.protection.sort = False
        sheet.protection.selectLockedCells = True
        sheet.protection.selectUnlockedCells = False
